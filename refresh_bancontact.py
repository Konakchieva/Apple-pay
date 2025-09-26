#!/usr/bin/env python3
"""
Refresh 'applepay_rep_perf_BANCONTACT_WIRED_hidden.xlsx' from CSVs and ALSO create
'applepay_rep_perf_BANCONTACT_READY.xlsx' (no Data_* tabs, values only, no formulas).

CSV names (same folder) — either:
  Generic:      Metrics.csv, Decline.csv, Usage.csv, Merchant.csv
  OR Bancontact: BANCONTACT_metrics*.csv, BANCONTACT_DECLINE*.csv, BANCONTACT_USAGE*.csv,
                 BANCONTACT_merchant*.csv

Run:
  py refresh_bancontact.py

Note: No Reporting-Month logic in this version. Circular-reference safe.
"""

import sys, math
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent

# Adjust if your workbook name differs:
XLSX  = HERE / "applepay_rep_perf_BANCONTACT_WIRED_hidden.xlsx"
READY = HERE / "applepay_rep_perf_BANCONTACT_READY.xlsx"

CSV_CANDIDATES = {
    "Data_Metrics":   ["Metrics.csv", "BANCONTACT_metrics*.csv", "*metrics*.csv"],
    "Data_Declines":  ["Decline.csv", "BANCONTACT_DECLINE*.csv", "*decline*.csv"],
    "Data_Usage":     ["Usage.csv", "BANCONTACT_USAGE*.csv", "*usage*.csv"],
    "Data_Merchant":  ["Merchant.csv", "BANCONTACT_merchant*.csv", "*merchant*.csv"],
}

# ---------------- helpers ----------------
def resolve_csv(patterns):
    for pat in patterns:
        for p in HERE.glob(pat):
            if p.is_file():
                return p
    return None

def read_csv_robust(p: Path):
    try:
        return pd.read_csv(p)
    except UnicodeDecodeError:
        return pd.read_csv(p, encoding="latin1")

def write_dataframe_to_sheet(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    for i, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    ws.sheet_state = "hidden"
    print(f"Updated {sheet_name}: {len(df)} rows × {len(df.columns)} cols")

def _anchor_cell(ws, cell):
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell

def to_num(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return 0.0
        return float(v)
    s = str(v).strip()
    if s == "": return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg: s = s[1:-1]
    pct = "%" in s
    s = (s.replace("%","").replace("€","").replace("\u00A0"," ").replace(" ",""))
    if s.count(",") > 0 and s.count(".") == 0: s = s.replace(",", ".")
    else: s = s.replace(",", "")
    try:
        x = float(s)
        if neg: x = -x
        if pct: x /= 100.0
        if math.isnan(x) or math.isinf(x): return 0.0
        return x
    except Exception:
        return 0.0

def data_nrows(wb, wsname, first_col=1):
    if wsname not in wb.sheetnames: return 0
    ws = wb[wsname]
    r = 2; cnt = 0
    while r <= ws.max_row and ws.cell(row=r, column=first_col).value not in (None, ""):
        cnt += 1; r += 1
    return cnt

def find_label_neighbor(ws, label_text: str):
    target = label_text.lower()
    max_r = min(ws.max_row, 60)
    max_c = min(ws.max_column, 30)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and target in v.lower():
                return (r, c+1, ws.cell(row=r, column=c+1).coordinate)
    return None

# ------------- wire visible sheets (no Fraud) -------------
def ensure_report_formulas_and_cosmetics(wb):
    # Metrics
    if "Metrics" in wb.sheetnames and "Data_Metrics" in wb.sheetnames:
        ws = wb["Metrics"]
        ws["B7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E7"].value  = '=SUM(B7:D7)'
        ws["B8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_PP",Data_Metrics!$1:$1,0)),IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_POS_PP",Data_Metrics!$1:$1,0)),""))'
        ws["E8"].value  = '=SUM(B8:D8)'
        ws["B9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E9"].value  = '=(B7*B9 + C7*C9 + D7*D9) / IF(E7=0,1,E7)'
        ws["B10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E10"].value = '=1 - E9'
        ws["B11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E11"].value = '=(B8*B11 + C8*C11 + D8*D11) / IF(E8=0,1,E8)'
        ws["B12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E12"].value = '=1 - E11'
        ws["B14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E14"].value = '=SUM(B14:D14)'
        # IMPORTANT: do NOT auto-fill the neighbor of 'Monthly DPAN transaction count' here (prevents circular refs)

    # Declines + cosmetics
    if "Declines" in wb.sheetnames and "Data_Declines" in wb.sheetnames:
        ws = wb["Declines"]
        n = data_nrows(wb, "Data_Declines")
        for i in range(n):
            r = 8 + i; off = i + 2
            ws[f"A{r}"].value = f'=IFERROR(INDEX(Data_Declines!$A:$A,{off}),"")'
            ws[f"B{r}"].value = f'=IFERROR(INDEX(Data_Declines!$B:$B,{off}),"")'
            ws[f"C{r}"].value = f'=IFERROR(INDEX(Data_Declines!$C:$C,{off}),"")'
            ws[f"D{r}"].value = f'=IFERROR(INDEX(Data_Declines!$D:$D,{off}),"")'
            ws[f"F{r}"].value = f'=IFERROR(INDEX(Data_Declines!$E:$E,{off}),"")'
            ws[f"G{r}"].value = f'=IFERROR(INDEX(Data_Declines!$F:$F,{off}),"")'
            ws[f"H{r}"].value = f'=IFERROR(INDEX(Data_Declines!$G:$G,{off}),"")'
        labels = ["Transaction Size","< 10€","€10 - €25","€25 - €50","€50 - €100","€100 - €250","€250 - €1000",">= €1000","Total"]
        for i, text in enumerate(labels, start=7):
            ws[f"A{i}"].value = text
        ws["A17"].value = "* Leave cell blank if not applicable"
        ws["B15"].value = "=SUM(B8:B14)"; ws["C15"].value = "=SUM(C8:C14)"
        ws["F15"].value = "=SUM(F8:F14)"; ws["G15"].value = "=SUM(G8:G14)"

    # Usage Frequency
    if "Usage Frequency" in wb.sheetnames and "Data_Usage" in wb.sheetnames:
        ws = wb["Usage Frequency"]
        for r in range(8,19):
            ws[f"F{r}"].value = f'=IFERROR(INDEX(Data_Usage!$1:$2,2,MATCH($B{r},Data_Usage!$1:$1,0)),"")'
            ws[f"G{r}"].value = f'=IFERROR(IF($F{r}=0,0,$F{r}/$F$19),"")'
        ws["F19"].value = "=SUM(F8:F18)"

    # Merchant Report (header-safe)
    if "Merchant Report" in wb.sheetnames and "Data_Merchant" in wb.sheetnames:
        ws = wb["Merchant Report"]
        for i in range(100):
            r = 7 + i; off = i + 2
            ws[f"A{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$1:$1048576,{off},MATCH("RANK",Data_Merchant!$1:$1,0)),"")'
            ws[f"B{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$1:$1048576,{off},MATCH("NOM_CMR",Data_Merchant!$1:$1,0)),"")'
            ws[f"C{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$1:$1048576,{off},MATCH("PERC",Data_Merchant!$1:$1,0)),"")'
            ws[f"D{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$1:$1048576,{off},MATCH("SPENT",Data_Merchant!$1:$1,0)),"")'
            ws[f"E{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$1:$1048576,{off},MATCH("CNT",Data_Merchant!$1:$1,0)),"")'

# ------------- READY (values only) -------------
def create_ready_values_only(wb_src):
    def df_from_sheet(name):
        if name not in wb_src.sheetnames: return None
        ws = wb_src[name]
        headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        rows = []
        r=2
        while r<=ws.max_row and ws.cell(row=r, column=1).value not in (None,""):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, len(headers)+1)])
            r+=1
        return pd.DataFrame(rows, columns=headers) if rows else pd.DataFrame(columns=headers)

    dm = df_from_sheet("Data_Metrics")
    dd = df_from_sheet("Data_Declines")
    du = df_from_sheet("Data_Usage")
    dmerch = df_from_sheet("Data_Merchant")

    def set_val(ws, addr, val): ws[addr].value = None if val is None else val

    # Metrics (values)
    if "Metrics" in wb_src.sheetnames and dm is not None and not dm.empty:
        ws = wb_src["Metrics"]
        row = dm.iloc[0].to_dict()
        b7 = to_num(row.get("CNT_DPAN_DEBIT", 0))
        c7 = to_num(row.get("CNT_DPAN_CREDIT", 0))
        d7 = to_num(row.get("CNT_DPAN_PP", row.get("CNT_DPAN_POS_PP", 0)))
        e7 = b7 + c7 + d7
        for a, v in [("B7", b7), ("C7", c7), ("D7", d7), ("E7", e7)]: set_val(ws, a, v)
        b8 = to_num(row.get("SUM_EXP_DPAN_DEBIT", 0))
        c8 = to_num(row.get("SUM_EXP_DPAN_CREDIT", 0))
        d8 = to_num(row.get("SUM_EXP_DPAN_PP", row.get("SUM_EXP_DPAN_POS_PP", 0)))
        e8 = b8 + c8 + d8
        for a, v in [("B8", b8), ("C8", c8), ("D8", d8), ("E8", e8)]: set_val(ws, a, v)
        b9 = to_num(row.get("PERC_DPAN_POS_DEBIT", 0))
        c9 = to_num(row.get("PERC_DPAN_POS_CREDIT", 0))
        d9 = to_num(row.get("PERC_DPAN_POS_PP", 0))
        e9 = (b7*b9 + c7*c9 + d7*d9) / (e7 or 1.0)
        for a, v in [("B9", b9), ("C9", c9), ("D9", d9), ("E9", e9)]: set_val(ws, a, v)
        b10 = to_num(row.get("PERC_DPAN_REM_DEBIT", 0))
        c10 = to_num(row.get("PERC_DPAN_REM_CREDIT", 0))
        d10 = to_num(row.get("PERC_DPAN_REM_PP", 0))
        e10 = 1.0 - e9
        for a, v in [("B10", b10), ("C10", c10), ("D10", d10), ("E10", e10)]: set_val(ws, a, v)
        b11 = to_num(row.get("PERC_EXP_DPAN_POS_DEBIT", 0))
        c11 = to_num(row.get("PERC_EXP_DPAN_POS_CREDIT", 0))
        d11 = to_num(row.get("PERC_EXP_DPAN_POS_PP", 0))
        e11 = (b8*b11 + c8*c11 + d8*d11) / (e8 or 1.0)
        for a, v in [("B11", b11), ("C11", c11), ("D11", d11), ("E11", e11)]: set_val(ws, a, v)
        b12 = to_num(row.get("PERC_EXP_DPAN_REM_DEBIT", 0))
        c12 = to_num(row.get("PERC_EXP_DPAN_REM_CREDIT", 0))
        d12 = to_num(row.get("PERC_EXP_DPAN_REM_PP", 0))
        e12 = 1.0 - e11
        for a, v in [("B12", b12), ("C12", c12), ("D12", d12), ("E12", e12)]: set_val(ws, a, v)
        b14 = to_num(row.get("CNT_ACTIVE_DPAN_DEBIT", 0))
        c14 = to_num(row.get("CNT_ACTIVE_DPAN_CREDIT", 0))
        d14 = to_num(row.get("CNT_ACTIVE_DPAN_PP", 0))
        e14 = b14 + c14 + d14
        for a, v in [("B14", b14), ("C14", c14), ("D14", d14), ("E14", e14)]: set_val(ws, a, v)

        # Fill the standalone label's neighbor only in READY (values-only), and only if safe
        nb = find_label_neighbor(ws, "Monthly DPAN transaction count")
        if nb:
            _, _, addr = nb
            if addr not in {"B7","C7","D7","E7"}:
                set_val(ws, addr, e7)  # write numeric value only

    # Declines (values + totals)
    if "Declines" in wb_src.sheetnames and dd is not None and not dd.empty:
        ws = wb_src["Declines"]
        n = len(dd)
        for i in range(n):
            r = 8 + i
            for (col, idx) in [("B",1),("C",2),("D",3),("F",4),("G",5),("H",6)]:
                v = dd.iloc[i, idx] if idx < dd.shape[1] else None
                set_val(ws, f"{col}{r}", to_num(v))
        def sum_col(c):
            return sum(to_num(ws[f"{c}{rr}"].value) for rr in range(8,15))
        set_val(ws,"B15",sum_col("B")); set_val(ws,"C15",sum_col("C"))
        set_val(ws,"F15",sum_col("F")); set_val(ws,"G15",sum_col("G"))

    # Usage Frequency (values)
    if "Usage Frequency" in wb_src.sheetnames and du is not None and not du.empty:
        ws = wb_src["Usage Frequency"]
        vals = du.iloc[0].to_dict()
        for r in range(8,19):
            set_val(ws, f"F{r}", to_num(vals.get(ws[f"B{r}"].value, 0)))
        total = sum(to_num(ws[f"F{r}"].value) for r in range(8,19))
        set_val(ws, "F19", total)
        for r in range(8,19):
            fv = to_num(ws[f"F{r}"].value)
            set_val(ws, f"G{r}", (fv/total) if total else 0.0)

    # Merchant (values as-is)
    if "Merchant Report" in wb_src.sheetnames and dmerch is not None and not dmerch.empty:
        ws = wb_src["Merchant Report"]
        for i in range(min(100, len(dmerch))):
            r = 7 + i
            for j, col in enumerate(["RANK","NOM_CMR","PERC","SPENT","CNT"], start=1):
                if col in dmerch.columns:
                    ws[f"{'ABCDE'[j-1]}{r}"].value = dmerch.iloc[i][col]

    # Drop Data_* tabs
    for name in list(wb_src.sheetnames):
        if name.startswith("Data_"):
            wb_src.remove(wb_src[name])

    wb_src.save(READY)
    print("Ready file:", READY.name)

# ---------------- main ----------------
def main():
    # Resolve CSVs
    resolved = {}
    missing = []
    for sheet, pats in CSV_CANDIDATES.items():
        p = resolve_csv(pats)
        if p is None: missing.append(sheet)
        else: resolved[sheet] = p
    if missing:
        print("ERROR: Missing CSVs for:", ", ".join(missing))
        sys.exit(1)
    if not XLSX.exists():
        print(f"ERROR: Excel file not found: {XLSX.name}")
        sys.exit(1)

    try:
        wb = load_workbook(XLSX, data_only=False, keep_links=True)
    except Exception as e:
        print("ERROR: Failed to open workbook:", e); sys.exit(1)

    for sheet, path in resolved.items():
        df = read_csv_robust(path)
        write_dataframe_to_sheet(wb, sheet, df)

    ensure_report_formulas_and_cosmetics(wb)

    try:
        wb.save(XLSX)
        print("✅ Refreshed:", XLSX.name)
    except PermissionError:
        print(f"ERROR: Close '{XLSX.name}' in Excel and run again."); sys.exit(1)

    wb_ready = load_workbook(XLSX, data_only=False, keep_links=True)
    create_ready_values_only(wb_ready)

if __name__ == "__main__":
    main()
