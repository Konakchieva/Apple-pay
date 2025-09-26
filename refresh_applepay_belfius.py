#!/usr/bin/env python3
"""
Refresh a BELFIUS Apple Pay workbook from CSVs and create a values-only copy.

- WIRED workbook: updates hidden Data_* sheets + rewires formulas on visible tabs.
- READY workbook: values only (no formulas, no Data_* sheets). Safe to send out.
- Reporting Month: stamped on EVERY run as the previous month (Month YYYY).
  * Applies to all visible sheets EXCEPT those whose title contains 'Glossary'.
  * If a 'Reporting Month' label is found, it is updated.
  * If not found, 'Reporting Month: <Month YYYY>' is written to A2.

CSV names required (first match wins):
  Metrics.csv, Decline.csv, Fraud.csv, Usage.csv, Merchant.csv
(You can also pass a CSV folder; the script will pick the first matching file.)

No circular references (we never write a helper next to the “Monthly DPAN transaction count” label in WIRED).

Usage examples
--------------
# Default names in current folder
py refresh_belfius.py

# Custom locations
py refresh_belfius.py --workbook "applepay_rep_perf_BELFIUS_DATAWIRED.xlsx" \
                      --csv-dir "." \
                      --ready-out "applepay_rep_perf_BELFIUS_READY.xlsx" \
                      -v
"""

from __future__ import annotations
import sys
import math
import argparse
import logging
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------- configuration ----------------

CSV_PATTERNS: Dict[str, List[str]] = {
    "Data_Metrics":   ["Metrics.csv", "*metrics*.csv"],
    "Data_Declines":  ["Decline.csv", "*decline*.csv"],
    "Data_Fraud":     ["Fraud.csv", "*fraud*.csv"],
    "Data_Usage":     ["Usage.csv", "*usage*.csv"],
    "Data_Merchant":  ["Merchant.csv", "*merchant*.csv"],
}

DEFAULT_WIRED = "applepay_rep_perf_BELFIUS_DATAWIRED.xlsx"
DEFAULT_READY = "applepay_rep_perf_BELFIUS_READY.xlsx"

# Sheets that should NOT receive a Reporting Month stamp
EXCLUDE_REPORTING_MONTH_TOKENS = ("glossary",)  # case-insensitive substring match

# ---------------- logging ----------------

def setup_logging(verbose: bool) -> None:
    lvl = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=lvl,
        format="%(asctime)s %(levelname)s: %(message)s",
        datefmt="%H:%M:%S",
    )

# ---------------- helpers ----------------

def prev_month_label() -> str:
    """Return previous month as 'Month YYYY' (based on local system date)."""
    first = date.today().replace(day=1)
    last_prev = first - timedelta(days=1)
    return last_prev.strftime("%B %Y")

def resolve_csvs(csv_dir: Path) -> Dict[str, Path]:
    found: Dict[str, Path] = {}
    for sheet, patterns in CSV_PATTERNS.items():
        match: Optional[Path] = None
        for pat in patterns:
            for p in csv_dir.glob(pat):
                if p.is_file():
                    match = p
                    break
            if match:
                break
        if not match:
            logging.error("Missing CSV for %s (looked for: %s)", sheet, ", ".join(patterns))
            continue
        found[sheet] = match
        logging.info("CSV → %-15s %s", sheet, match.name)
    return found

def read_csv_robust(p: Path) -> pd.DataFrame:
    try:
        df = pd.read_csv(p)
    except UnicodeDecodeError:
        df = pd.read_csv(p, encoding="latin1")
    return df

def write_dataframe_to_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    # headers
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))

    # data rows
    for i, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    ws.sheet_state = "hidden"
    logging.info("Wrote %-15s %5d rows × %d cols (hidden)", sheet_name, len(df), len(df.columns))

def _anchor_cell(ws: Worksheet, cell):
    # return the top-left cell of a merged range if 'cell' is inside one
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell

def to_num(v) -> float:
    """Coerce strings like '1 234,56', '€1,234.56', '(123)', '12%' → float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return 0.0
        return float(v)
    s = str(v).strip()
    if s == "":
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg: s = s[1:-1]
    pct = "%" in s
    s = s.replace("%", "").replace("€", "").replace("\u00A0", " ").replace(" ", "")
    if s.count(",") > 0 and s.count(".") == 0:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        x = float(s)
        if neg: x = -x
        if pct: x /= 100.0
        if math.isnan(x) or math.isinf(x): return 0.0
        return x
    except Exception:
        return 0.0

def data_nrows(wb, wsname: str, first_col: int = 1) -> int:
    if wsname not in wb.sheetnames:
        return 0
    ws = wb[wsname]
    r = 2
    cnt = 0
    while r <= ws.max_row and ws.cell(row=r, column=first_col).value not in (None, ""):
        cnt += 1
        r += 1
    return cnt

def find_label_neighbor(ws: Worksheet, label_text: str) -> Optional[Tuple[int,int,str]]:
    """Find the cell to the right of a label text (case-insensitive)."""
    target = label_text.lower()
    max_r = min(ws.max_row, 60)
    max_c = min(ws.max_column, 30)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and target in v.lower():
                addr = ws.cell(row=r, column=c+1).coordinate
                return (r, c+1, addr)
    return None

# ---------------- Reporting Month (safe) ----------------

def upsert_reporting_month(ws: Worksheet, month_label: str) -> None:
    """
    Update or create the 'Reporting Month' on a worksheet.
    - If any cell contains 'Reporting Month' (case-insensitive), update it:
        * If the label and month are in the SAME cell (e.g., 'Reporting Month: August 2025'),
          replace the entire label with 'Reporting Month: <Month YYYY>'.
        * Else, write <Month YYYY> into the cell to the right.
    - If not found, write 'Reporting Month: <Month YYYY>' to A2.
    Merged-cell safe: always writes to the anchor cell; never clears neighbors.
    """
    max_r = min(ws.max_row, 10)
    max_c = min(ws.max_column, 30)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and "reporting month" in val.lower():
                anchor = _anchor_cell(ws, cell)
                text = val.strip()
                if ":" in text:
                    anchor.value = f"Reporting Month: {month_label}"
                else:
                    right = ws.cell(row=r, column=c + 1)
                    _anchor_cell(ws, right).value = month_label
                return
    a2 = _anchor_cell(ws, ws["A2"])
    a2.value = f"Reporting Month: {month_label}"

def set_reporting_month_on_workbook(wb) -> None:
    label = prev_month_label()
    for ws in wb.worksheets:
        name = ws.title
        lname = name.lower()
        if lname.startswith("data_"):
            continue
        if any(tok in lname for tok in EXCLUDE_REPORTING_MONTH_TOKENS):
            continue
        upsert_reporting_month(ws, label)

# ---------------- wire visible sheets ----------------

def wire_visible_sheets(wb) -> None:
    # Metrics
    if "Metrics" in wb.sheetnames and "Data_Metrics" in wb.sheetnames:
        ws = wb["Metrics"]
        ws["B7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E7"].value  = '=SUM(B7:D7)'
        ws["B8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_PP",Data_Metrics!$1:$1,0)),IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_POS_PP",Data_Metrics!$1:$1,0)),""))'
        ws["E8"].value  = '=SUM(B8:D8)'
        ws["B9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E9"].value  = '=(B7*B9 + C7*C9 + D7*D9) / IF(E7=0,1,E7)'
        ws["B10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E10"].value = '=1 - E9'
        ws["B11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E11"].value = '=(B8*B11 + C8*C11 + D8*D11) / IF(E8=0,1,E8)'
        ws["B12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E12"].value = '=1 - E11'
        ws["B14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E14"].value = '=SUM(B14:D14)'
        # No helper next to "Monthly DPAN transaction count" in WIRED (avoid circular refs)

    # Declines + cosmetics (A..D and F..H mapped; labels + totals)
    if "Declines" in wb.sheetnames and "Data_Declines" in wb.sheetnames:
        ws = wb["Declines"]
        n = data_nrows(wb, "Data_Declines")
        for i in range(n):
            r = 8 + i; off = i + 2
            ws[f"A{r}"].value = f'=IFERROR(INDEX(Data_Declines!$A:$A,{off}),"")'
            ws[f"B{r}"].value = f'=IFERROR(INDEX(Data_Declines!$B:$B,{off}),"")'
            ws[f"C{r}"].value = f'=IFERROR(INDEX(Data_Declines!$C:$C,{off}),"")'
            ws[f"D{r}"].value = f'=IFERROR(INDEX(Data_Declines!$D:$D,{off}),"")'
            ws[f"F{r}"].value = f'=IFERROR(INDEX(Data_Declines!$E:$E,{off}),"")'
            ws[f"G{r}"].value = f'=IFERROR(INDEX(Data_Declines!$F:$F,{off}),"")'
            ws[f"H{r}"].value = f'=IFERROR(INDEX(Data_Declines!$G:$G,{off}),"")'
        labels = ["Transaction Size","< 10€","€10 - €25","€25 - €50","€50 - €100","€100 - €250","€250 - €1000",">= €1000","Total"]
        for i, text in enumerate(labels, start=7):
            ws[f"A{i}"].value = text
        ws["A17"].value = "* Leave cell blank if not applicable"
        ws["B15"].value = "=SUM(B8:B14)"; ws["C15"].value = "=SUM(C8:C14)"
        ws["F15"].value = "=SUM(F8:F14)"; ws["G15"].value = "=SUM(G8:G14)"

    # Usage Frequency
    if "Usage Frequency" in wb.sheetnames and "Data_Usage" in wb.sheetnames:
        ws = wb["Usage Frequency"]
        for r in range(8,19):
            ws[f"F{r}"].value = f'=IFERROR(INDEX(Data_Usage!$1:$2,2,MATCH($B{r},Data_Usage!$1:$1,0)),"")'
            ws[f"G{r}"].value = f'=IFERROR(IF($F{r}=0,0,$F{r}/$F$19),"")'
        ws["F19"].value = "=SUM(F8:F18)"

    # Merchant Report
    if "Merchant Report" in wb.sheetnames and "Data_Merchant" in wb.sheetnames:
        ws = wb["Merchant Report"]
        for i in range(100):
            r = 7 + i; off = i + 2
            ws[f"A{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$A:$A,{off}),"")'
            ws[f"B{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$B:$B,{off}),"")'
            ws[f"C{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$C:$C,{off}),"")'
            ws[f"D{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$D:$D,{off}),"")'
            ws[f"E{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$E:$E,{off}),"")'

    # Fraud (Devices→DPANs header tolerance)
    if "Fraud" in wb.sheetnames and "Data_Fraud" in wb.sheetnames:
        ws = wb["Fraud"]
        def fraud_formula(col: str, r: int) -> str:
            return (
                f'=IF(OR($A{r}="",ISNUMBER(SEARCH("Leave cell blank",$A{r}))),"",'
                f'IFERROR(INDEX(Data_Fraud!${col}:${col},'
                f'MATCH(SUBSTITUTE($A{r},"Devices","DPANs"),Data_Fraud!$A:$A,0)),""))'
            )
        for r in range(7, 47):
            ws[f"B{r}"].value = fraud_formula("B", r)
            ws[f"C{r}"].value = fraud_formula("C", r)
            ws[f"D{r}"].value = fraud_formula("D", r)
            ws[f"E{r}"].value = fraud_formula("E", r)

# ---------------- READY (values only) ----------------

def create_ready_values_only(wb_src, ready_path: Path) -> None:
    """Write values into visible sheets, drop Data_* tabs, stamp Reporting Month, and save READY."""
    def df_from_sheet(name: str) -> Optional[pd.DataFrame]:
        if name not in wb_src.sheetnames:
            return None
        ws = wb_src[name]
        headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        rows = []
        r = 2
        while r <= ws.max_row and ws.cell(row=r, column=1).value not in (None, ""):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, len(headers)+1)])
            r += 1
        return pd.DataFrame(rows, columns=headers) if rows else pd.DataFrame(columns=headers)

    dm = df_from_sheet("Data_Metrics")
    dd = df_from_sheet("Data_Declines")
    du = df_from_sheet("Data_Usage")
    dmerch = df_from_sheet("Data_Merchant")
    dfraud = df_from_sheet("Data_Fraud")

    def set_val(ws: Worksheet, addr: str, val):
        _anchor_cell(ws, ws[addr]).value = None if val is None else val

    # Metrics (values)
    if "Metrics" in wb_src.sheetnames and dm is not None and not dm.empty:
        ws = wb_src["Metrics"]
        row = dm.iloc[0].to_dict()
        b7 = to_num(row.get("CNT_DPAN_DEBIT", 0))
        c7 = to_num(row.get("CNT_DPAN_CREDIT", 0))
        d7 = to_num(row.get("CNT_DPAN_PP", row.get("CNT_DPAN_POS_PP", 0)))
        e7 = b7 + c7 + d7
        for a, v in [("B7", b7), ("C7", c7), ("D7", d7), ("E7", e7)]: set_val(ws, a, v)

        b8 = to_num(row.get("SUM_EXP_DPAN_DEBIT", 0))
        c8 = to_num(row.get("SUM_EXP_DPAN_CREDIT", 0))
        d8 = to_num(row.get("SUM_EXP_DPAN_PP", row.get("SUM_EXP_DPAN_POS_PP", 0)))
        e8 = b8 + c8 + d8
        for a, v in [("B8", b8), ("C8", c8), ("D8", d8), ("E8", e8)]: set_val(ws, a, v)

        b9 = to_num(row.get("PERC_DPAN_POS_DEBIT", 0))
        c9 = to_num(row.get("PERC_DPAN_POS_CREDIT", 0))
        d9 = to_num(row.get("PERC_DPAN_POS_PP", 0))
        e9 = (b7*b9 + c7*c9 + d7*d9) / (e7 or 1.0)
        for a, v in [("B9", b9), ("C9", c9), ("D9", d9), ("E9", e9)]: set_val(ws, a, v)

        b10 = to_num(row.get("PERC_DPAN_REM_DEBIT", 0))
        c10 = to_num(row.get("PERC_DPAN_REM_CREDIT", 0))
        d10 = to_num(row.get("PERC_DPAN_REM_PP", 0))
        e10 = 1.0 - e9
        for a, v in [("B10", b10), ("C10", c10), ("D10", d10), ("E10", e10)]: set_val(ws, a, v)

        b11 = to_num(row.get("PERC_EXP_DPAN_POS_DEBIT", 0))
        c11 = to_num(row.get("PERC_EXP_DPAN_POS_CREDIT", 0))
        d11 = to_num(row.get("PERC_EXP_DPAN_POS_PP", 0))
        e11 = (b8*b11 + c8*c11 + d8*d11) / (e8 or 1.0)
        for a, v in [("B11", b11), ("C11", c11), ("D11", d11), ("E11", e11)]: set_val(ws, a, v)

        b12 = to_num(row.get("PERC_EXP_DPAN_REM_DEBIT", 0))
        c12 = to_num(row.get("PERC_EXP_DPAN_REM_CREDIT", 0))
        d12 = to_num(row.get("PERC_EXP_DPAN_REM_PP", 0))
        e12 = 1.0 - e11
        for a, v in [("B12", b12), ("C12", c12), ("D12", d12), ("E12", e12)]: set_val(ws, a, v)

        b14 = to_num(row.get("CNT_ACTIVE_DPAN_DEBIT", 0))
        c14 = to_num(row.get("CNT_ACTIVE_DPAN_CREDIT", 0))
        d14 = to_num(row.get("CNT_ACTIVE_DPAN_PP", 0))
        e14 = b14 + c14 + d14
        for a, v in [("B14", b14), ("C14", c14), ("D14", d14), ("E14", e14)]: set_val(ws, a, v)

    # Declines (values + totals)
    if "Declines" in wb_src.sheetnames and dd is not None and not dd.empty:
        ws = wb_src["Declines"]
        n = len(dd)
        for i in range(n):
            r = 8 + i
            for (col, idx) in [("B",1),("C",2),("D",3),("F",4),("G",5),("H",6)]:
                v = dd.iloc[i, idx] if idx < dd.shape[1] else None
                set_val(ws, f"{col}{r}", to_num(v))
        def sum_col(c: str) -> float:
            return sum(to_num(ws[f"{c}{rr}"].value) for rr in range(8, 15))
        set_val(ws, "B15", sum_col("B")); set_val(ws, "C15", sum_col("C"))
        set_val(ws, "F15", sum_col("F")); set_val(ws, "G15", sum_col("G"))

    # Usage Frequency (values)
    if "Usage Frequency" in wb_src.sheetnames and du is not None and not du.empty:
        ws = wb_src["Usage Frequency"]
        vals = du.iloc[0].to_dict()
        for r in range(8,19):
            ws[f"F{r}"].value = to_num(vals.get(ws[f"B{r}"].value, 0))
        total = sum(to_num(ws[f"F{r}"].value) for r in range(8,19))
        ws["F19"].value = total
        for r in range(8,19):
            fv = to_num(ws[f"F{r}"].value)
            ws[f"G{r}"].value = (fv / total) if total else 0.0

    # Merchant (values)
    if "Merchant Report" in wb_src.sheetnames and dmerch is not None and not dmerch.empty:
        ws = wb_src["Merchant Report"]
        for i in range(min(100, len(dmerch))):
            r = 7 + i
            for j, col in enumerate(["RANK","NOM_CMR","PERC","SPENT","CNT"], start=1):
                if col in dmerch.columns:
                    ws[f"{'ABCDE'[j-1]}{r}"].value = dmerch.iloc[i][col]

    # Fraud (values)
    if "Fraud" in wb_src.sheetnames and dfraud is not None and not dfraud.empty:
        ws = wb_src["Fraud"]
        def norm(s): 
            return " ".join(str(s).replace("\xa0"," ").replace("\r"," ").replace("\n"," ").split()) if s is not None else ""
        lookup = {norm(dfraud.iloc[i,0]).replace("Devices","DPANs"): dfraud.iloc[i].to_dict() for i in range(len(dfraud))}
        for r in range(7,47):
            label = ws[f"A{r}"].value
            if not label or "Leave cell blank" in str(label):
                for col in "BCDE": set_val(ws,f"{col}{r}",None)
                continue
            rec = lookup.get(norm(label).replace("Devices","DPANs"))
            if rec:
                for col, key in zip("BCDE", ["Debit","Credit","Prepaid","Total"]):
                    ws[f"{col}{r}"].value = to_num(rec.get(key))

    # Drop Data_* tabs
    for name in list(wb_src.sheetnames):
        if name.startswith("Data_"):
            wb_src.remove(wb_src[name])

    # Stamp Reporting Month on READY as well
    set_reporting_month_on_workbook(wb_src)

    wb_src.save(ready_path)
    logging.info("✅ Saved READY: %s", ready_path.name)

# ---------------- CLI / main ----------------

def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="BELFIUS Apple Pay workbook refresher")
    ap.add_argument("--workbook", "-w", default=DEFAULT_WIRED, help="Path to WIRED workbook (xlsx)")
    ap.add_argument("--csv-dir", "-c", default=".", help="Folder with CSVs")
    ap.add_argument("--ready-out", "-o", default=DEFAULT_READY, help="Output READY workbook filename")
    ap.add_argument("--no-ready", action="store_true", help="Skip generating READY workbook")
    ap.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")
    return ap.parse_args()

def main() -> int:
    args = parse_args()
    setup_logging(args.verbose)

    csv_dir = Path(args.csv_dir).resolve()
    wired_path = Path(args.workbook).resolve()
    ready_path = Path(args.ready_out).resolve()

    logging.info("Workbook: %s", wired_path)
    logging.info("CSV dir : %s", csv_dir)

    resolved = resolve_csvs(csv_dir)
    missing_keys = [k for k in CSV_PATTERNS.keys() if k not in resolved]
    if missing_keys:
        logging.error("Aborting: missing CSVs for %s", ", ".join(missing_keys))
        return 1

    if not wired_path.exists():
        logging.error("Workbook not found: %s", wired_path.name)
        return 1

    # Open WIRED workbook
    try:
        wb = load_workbook(wired_path, data_only=False, keep_links=True)
    except Exception as e:
        logging.error("Failed to open workbook: %s", e)
        return 2

    # Update Data_* tabs
    for sheet, path in resolved.items():
        df = read_csv_robust(path)
        write_dataframe_to_sheet(wb, sheet, df)

    # Wire formulas / cosmetics
    wire_visible_sheets(wb)

    # Stamp Reporting Month (previous month) on all report sheets except Glossary
    set_reporting_month_on_workbook(wb)

    # Save WIRED
    try:
        wb.save(wired_path)
    except PermissionError:
        logging.error("Close '%s' in Excel and run again.", wired_path.name)
        return 2
    logging.info("✅ Saved WIRED: %s", wired_path.name)

    # READY (values-only)
    if not args.no_ready:
        wb_ready = load_workbook(wired_path, data_only=False, keep_links=True)
        create_ready_values_only(wb_ready, ready_path)

    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print()
        sys.exit(130)
    except Exception as e:
        logging.exception("Unexpected error: %s", e)
        sys.exit(3)
