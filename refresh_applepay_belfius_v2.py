#!/usr/bin/env python3
"""
Refresh 'applepay_rep_perf_BELFIUS_DATAWIRED.xlsx' from CSVs and ALSO create
'applepay_rep_perf_BELFIUS_READY.xlsx' (no Data_* tabs, values only).

Expected CSV names (same folder):
  Metrics.csv, Decline.csv, Fraud.csv, Usage.csv, Merchant.csv

Run:
  py refresh_applepay_belfius.py
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE / "applepay_rep_perf_BELFIUS_DATAWIRED.xlsx"
READY = HERE / "applepay_rep_perf_BELFIUS_READY.xlsx"

CSV_MAP = {
    "Data_Metrics":   "Metrics.csv",
    "Data_Declines":  "Decline.csv",
    "Data_Fraud":     "Fraud.csv",
    "Data_Usage":     "Usage.csv",
    "Data_Merchant":  "Merchant.csv",
}

def read_csv_robust(p: Path):
    try: return pd.read_csv(p)
    except UnicodeDecodeError: return pd.read_csv(p, encoding="latin1")

def write_dataframe_to_sheet(wb, sheet_name: str, df):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 0: ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    for i, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    ws.sheet_state = "hidden"
    print(f"Updated {sheet_name}: {len(df)} rows × {len(df.columns)} cols")

    # --------------------- Merged cell-safe writer ---------------------
def _anchor_cell(ws, cell):
    """Return the top-left cell of a merged range if 'cell' is merged; else return 'cell'."""
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell

# --------------------- Reporting Month ---------------------
def prev_month_label() -> str:
    """Return 'Month YYYY' for the previous calendar month."""
    first_of_this_month = date.today().replace(day=1)
    last_of_prev = first_of_this_month - timedelta(days=1)
    return last_of_prev.strftime("%B %Y")  # e.g., 'September 2025'

def set_reporting_month(wb):
    """Set 'Reporting Month' on row 2. Works even if A2/B2 (or any cell in row 2) is merged."""
    label = prev_month_label()
    for ws in wb.worksheets:
        # scan row 2 for the label
        wrote = False
        for col in range(1, 31):  # A..AE, generous
            cell = ws.cell(row=2, column=col)
            text = (str(cell.value).strip() if cell.value is not None else "")
            low = text.lower()
            if "reporting month" in low:
                # If the label cell already includes the phrase, write full string there
                if ":" in text or col == 1:
                    _anchor_cell(ws, cell).value = f"Reporting Month: {label}"
                else:
                    # Otherwise put the month into the next cell (value cell)
                    next_cell = ws.cell(row=2, column=col + 1)
                    _anchor_cell(ws, next_cell).value = label
                wrote = True
                break
        if not wrote:
            # fallback: write full string into A2
            _anchor_cell(ws, ws["A2"]).value = f"Reporting Month: {label}"

def ensure_report_formulas_and_cosmetics(wb):
    # Metrics formulas (commas for XLSX)
    if "Metrics" in wb.sheetnames and "Data_Metrics" in wb.sheetnames:
        ws = wb["Metrics"]
        ws["B7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D7"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_DPAN_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E7"].value  = '=SUM(B7:D7)'
        ws["B8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D8"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_PP",Data_Metrics!$1:$1,0)),IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("SUM_EXP_DPAN_POS_PP",Data_Metrics!$1:$1,0)),""))'
        ws["E8"].value  = '=SUM(B8:D8)'
        ws["B9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D9"].value  = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_POS_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E9"].value  = '=(B7*B9 + C7*C9 + D7*D9) / IF(E7=0,1,E7)'
        ws["B10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D10"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_DPAN_REM_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E10"].value = '=1 - E9'
        ws["B11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D11"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_POS_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E11"].value = '=(B8*B11 + C8*C11 + D8*D11) / IF(E8=0,1,E8)'
        ws["B12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D12"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("PERC_EXP_DPAN_REM_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E12"].value = '=1 - E11'
        ws["B14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_DEBIT",Data_Metrics!$1:$1,0)),"")'
        ws["C14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_CREDIT",Data_Metrics!$1:$1,0)),"")'
        ws["D14"].value = '=IFERROR(INDEX(Data_Metrics!$1:$1048576,2,MATCH("CNT_ACTIVE_DPAN_PP",Data_Metrics!$1:$1,0)),"")'
        ws["E14"].value = '=SUM(B14:D14)'

    # Declines mapping (A..D and F..H) + cosmetics (labels + totals)
    def data_nrows(wsname, first_col=1):
        if wsname not in wb.sheetnames: return 0
        wsx = wb[wsname]; r=2; cnt=0
        while r<=wsx.max_row and wsx.cell(row=r, column=first_col).value not in (None,""):
            cnt+=1; r+=1
        return cnt

    if "Declines" in wb.sheetnames and "Data_Declines" in wb.sheetnames:
        ws = wb["Declines"]
        n = data_nrows("Data_Declines")
        for i in range(n):
            r = 8 + i; off = i + 2
            ws[f"A{r}"].value = f'=IFERROR(INDEX(Data_Declines!$A:$A,{off}),"")'
            ws[f"B{r}"].value = f'=IFERROR(INDEX(Data_Declines!$B:$B,{off}),"")'
            ws[f"C{r}"].value = f'=IFERROR(INDEX(Data_Declines!$C:$C,{off}),"")'
            ws[f"D{r}"].value = f'=IFERROR(INDEX(Data_Declines!$D:$D,{off}),"")'
            ws[f"F{r}"].value = f'=IFERROR(INDEX(Data_Declines!$E:$E,{off}),"")'
            ws[f"G{r}"].value = f'=IFERROR(INDEX(Data_Declines!$F:$F,{off}),"")'
            ws[f"H{r}"].value = f'=IFERROR(INDEX(Data_Declines!$G:$G,{off}),"")'
        # Cosmetic labels in A7..A15, note at A17, totals row 15
        labels = ["Transaction Size","< 10€","€10 - €25","€25 - €50","€50 - €100","€100 - €250","€250 - €1000",">= €1000","Total"]
        for i, text in enumerate(labels, start=7):
            ws[f"A{i}"].value = text
        ws["A17"].value = "* Leave cell blank if not applicable"
        ws["B15"].value = "=SUM(B8:B14)"; ws["C15"].value = "=SUM(C8:C14)"
        ws["F15"].value = "=SUM(F8:F14)"; ws["G15"].value = "=SUM(G8:G14)"

    # Usage Frequency
    if "Usage Frequency" in wb.sheetnames and "Data_Usage" in wb.sheetnames:
        ws = wb["Usage Frequency"]
        for r in range(8,19):
            ws[f"F{r}"].value = f'=IFERROR(INDEX(Data_Usage!$1:$2,2,MATCH($B{r},Data_Usage!$1:$1,0)),"")'
            ws[f"G{r}"].value = f'=IFERROR(IF($F{r}=0,0,$F{r}/$F$19),"")'
        ws["F19"].value = "=SUM(F8:F18)"

    # Merchant Report
    if "Merchant Report" in wb.sheetnames and "Data_Merchant" in wb.sheetnames:
        ws = wb["Merchant Report"]
        # fill top 100 rows
        for i in range(100):
            r = 7 + i; off = i + 2
            ws[f"A{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$A:$A,{off}),"")'
            ws[f"B{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$B:$B,{off}),"")'
            ws[f"C{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$C:$C,{off}),"")'
            ws[f"D{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$D:$D,{off}),"")'
            ws[f"E{r}"].value = f'=IFERROR(INDEX(Data_Merchant!$E:$E,{off}),"")'

    # Fraud (with Devices->DPANs)
    if "Fraud" in wb.sheetnames and "Data_Fraud" in wb.sheetnames:
        ws = wb["Fraud"]
        def formula(col, r):
            return (f'=IF(OR($A{r}="",ISNUMBER(SEARCH("Leave cell blank",$A{r}))),"",'
                    f'IFERROR(INDEX(Data_Fraud!${col}:${col},'
                    f'MATCH(SUBSTITUTE($A{r},"Devices","DPANs"),Data_Fraud!$A:$A,0)),""))')
        for r in range(7,47):
            ws[f"B{r}"].value = formula("B", r)
            ws[f"C{r}"].value = formula("C", r)
            ws[f"D{r}"].value = formula("D", r)
            ws[f"E{r}"].value = formula("E", r)

def create_ready_values_only(wb_src):
    """Create a copy with values only and no Data_* tabs."""
    from copy import deepcopy
    wbv = wb_src  # work in-place on a second load
    import pandas as pd

    def df_from_sheet(name):
        if name not in wbv.sheetnames: return None
        ws = wbv[name]
        headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        rows = []
        r=2
        while r<=ws.max_row and ws.cell(row=r, column=1).value not in (None,""):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, len(headers)+1)])
            r+=1
        return pd.DataFrame(rows, columns=headers) if rows else pd.DataFrame(columns=headers)

    dm = df_from_sheet("Data_Metrics")
    dd = df_from_sheet("Data_Declines")
    du = df_from_sheet("Data_Usage")
    dmerch = df_from_sheet("Data_Merchant")
    dfraud = df_from_sheet("Data_Fraud")

    def set_val(ws, addr, val): ws[addr].value = None if val is None else val
    def norm(s):
        if s is None: return ""
        return " ".join(str(s).replace("\xa0"," ").replace("\r"," ").replace("\n"," ").split())

    # Metrics
    if "Metrics" in wbv.sheetnames and dm is not None and not dm.empty:
        ws = wbv["Metrics"]
        row = dm.iloc[0].to_dict()
        b7 = row.get("CNT_DPAN_DEBIT",0); c7=row.get("CNT_DPAN_CREDIT",0); d7=row.get("CNT_DPAN_PP",row.get("CNT_DPAN_POS_PP",0)); e7=(b7 or 0)+(c7 or 0)+(d7 or 0)
        for a,v in [("B7",b7),("C7",c7),("D7",d7),("E7",e7)]: set_val(ws,a,v)
        b8=row.get("SUM_EXP_DPAN_DEBIT",0); c8=row.get("SUM_EXP_DPAN_CREDIT",0); d8=row.get("SUM_EXP_DPAN_PP",row.get("SUM_EXP_DPAN_POS_PP",0)); e8=(b8 or 0)+(c8 or 0)+(d8 or 0)
        for a,v in [("B8",b8),("C8",c8),("D8",d8),("E8",e8)]: set_val(ws,a,v)
        b9=row.get("PERC_DPAN_POS_DEBIT",0); c9=row.get("PERC_DPAN_POS_CREDIT",0); d9=row.get("PERC_DPAN_POS_PP",0); e9=((b7 or 0)*(b9 or 0)+(c7 or 0)*(c9 or 0)+(d7 or 0)*(d9 or 0))/(e7 or 1)
        for a,v in [("B9",b9),("C9",c9),("D9",d9),("E9",e9)]: set_val(ws,a,v)
        b10=row.get("PERC_DPAN_REM_DEBIT",0); c10=row.get("PERC_DPAN_REM_CREDIT",0); d10=row.get("PERC_DPAN_REM_PP",0); e10=1-e9
        for a,v in [("B10",b10),("C10",c10),("D10",d10),("E10",e10)]: set_val(ws,a,v)
        b11=row.get("PERC_EXP_DPAN_POS_DEBIT",0); c11=row.get("PERC_EXP_DPAN_POS_CREDIT",0); d11=row.get("PERC_EXP_DPAN_POS_PP",0); e11=((b8 or 0)*(b11 or 0)+(c8 or 0)*(c11 or 0)+(d8 or 0)*(d11 or 0))/(e8 or 1)
        for a,v in [("B11",b11),("C11",c11),("D11",d11),("E11",e11)]: set_val(ws,a,v)
        b12=row.get("PERC_EXP_DPAN_REM_DEBIT",0); c12=row.get("PERC_EXP_DPAN_REM_CREDIT",0); d12=row.get("PERC_EXP_DPAN_REM_PP",0); e12=1-e11
        for a,v in [("B12",b12),("C12",c12),("D12",d12),("E12",e12)]: set_val(ws,a,v)
        b14=row.get("CNT_ACTIVE_DPAN_DEBIT",0); c14=row.get("CNT_ACTIVE_DPAN_CREDIT",0); d14=row.get("CNT_ACTIVE_DPAN_PP",0); e14=(b14 or 0)+(c14 or 0)+(d14 or 0)
        for a,v in [("B14",b14),("C14",c14),("D14",d14),("E14",e14)]: set_val(ws,a,v)

    # Declines
    if "Declines" in wbv.sheetnames and dd is not None and not dd.empty:
        ws = wbv["Declines"]
        n=len(dd)
        for i in range(n):
            r=8+i
            vals=[dd.iloc[i,j] if j<dd.shape[1] else None for j in range(7)]
            for (col,idx) in [("B",1),("C",2),("D",3),("F",4),("G",5),("H",6)]:
                set_val(ws, f"{col}{r}", vals[idx] if idx<len(vals) else None)
        def sum_col(col):
            s=0
            for rr in range(8,15):
                v=ws[f"{col}{rr}"].value or 0
                try: s+=float(v)
                except: pass
            return s
        set_val(ws,"B15",sum_col("B")); set_val(ws,"C15",sum_col("C"))
        set_val(ws,"F15",sum_col("F")); set_val(ws,"G15",sum_col("G"))

    # Usage
    if "Usage Frequency" in wbv.sheetnames and du is not None and not du.empty:
        ws=wbv["Usage Frequency"]
        row_vals=du.iloc[0].to_dict()
        for r in range(8,19):
            v=row_vals.get(ws[f"B{r}"].value,0); set_val(ws,f"F{r}",v)
        total=sum(float(ws[f"F{r}"].value or 0) for r in range(8,19))
        set_val(ws,"F19",total)
        for r in range(8,19):
            fv=float(ws[f"F{r}"].value or 0)
            set_val(ws,f"G{r}", (fv/total) if total else 0)

    # Merchant
    if "Merchant Report" in wbv.sheetnames and dmerch is not None and not dmerch.empty:
        ws=wbv["Merchant Report"]
        for i in range(min(100,len(dmerch))):
            r=7+i
            for j,col in enumerate(["RANK","NOM_CMR","PERC","SPENT","CNT"], start=1):
                if col in dmerch.columns: set_val(ws, f"{'ABCDE'[j-1]}{r}", dmerch.iloc[i][col])

    # Fraud
    if "Fraud" in wbv.sheetnames and dfraud is not None and not dfraud.empty:
        ws=wbv["Fraud"]
        def norm(s): return " ".join(str(s).replace("\xa0"," ").replace("\r"," ").replace("\n"," ").split()) if s is not None else ""
        lookup={norm(dfraud.iloc[i,0]).replace("Devices","DPANs"): dfraud.iloc[i].to_dict() for i in range(len(dfraud))}
        for r in range(7,47):
            label=ws[f"A{r}"].value
            if not label or "Leave cell blank" in str(label): 
                for col in "BCDE": set_val(ws,f"{col}{r}",None)
                continue
            rec=lookup.get(norm(label).replace("Devices","DPANs"))
            if rec:
                for col, key in zip("BCDE", ["Debit","Credit","Prepaid","Total"]):
                    set_val(ws, f"{col}{r}", rec.get(key))

    # Drop Data_* tabs
    for name in list(wbv.sheetnames):
        if name.startswith("Data_"): wbv.remove(wbv[name])

    wbv.save(READY)
    print("Ready file:", READY.name)

def main():
    # Ensure CSVs present
    missing=[name for name in CSV_MAP.values() if not (HERE/name).exists()]
    if missing:
        print("ERROR: Missing CSV files:", ", ".join(missing)); sys.exit(1)
    if not XLSX.exists():
        print(f"ERROR: Excel file not found: {XLSX.name}"); sys.exit(1)

    wb = load_workbook(XLSX, data_only=False, keep_links=True)

    # Write Data_* sheets
    for sheet, csv_name in CSV_MAP.items():
        df = read_csv_robust(HERE / csv_name)
        write_dataframe_to_sheet(wb, sheet, df)

    # Wire report formulas + Declines cosmetics
    ensure_report_formulas_and_cosmetics(wb)

    # Save main
    try:
        wb.save(XLSX)
        print("✅ Refreshed:", XLSX.name)
    except PermissionError:
        print(f"ERROR: Close '{XLSX.name}' in Excel and run again."); sys.exit(1)

    # Create READY (values-only) copy
    wb_ready = load_workbook(XLSX, data_only=False, keep_links=True)
    create_ready_values_only(wb_ready)

if __name__ == "__main__":
    main()
